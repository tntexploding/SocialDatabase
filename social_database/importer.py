"""标准来源记录、xlsx 适配和 SQLite 非破坏合并。"""

from collections.abc import Sequence
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.orm import Session as SessionType

from .config import DB_PATH, RELATION_FIELDS, REQUIRED_COLUMNS, SOURCE_COLUMNS
from .models import (
    Group,
    ImportBatch,
    Member,
    MemberGroupInfo,
    RelationObservation,
    init_db,
)
from .output import safe_print
from .search_index import (
    get_search_index_state,
    rebuild_search_index,
)

Record = dict[str, str | None]
SEARCH_CONTENT_RELATION_FIELDS = (
    "nickname",
    "card",
    "join_time",
    "last_sent_time",
    "title",
)


@dataclass(frozen=True)
class ParsedRecords:
    """标准来源记录及数据质量统计。"""

    rows: tuple[Record, ...]
    source_rows: int
    skipped_rows: int
    missing_user_id_rows: int
    missing_group_id_rows: int


# 保留 0.5.0 中可导入的类型名称，避免破坏已有 Python 调用方。
ParsedWorkbook = ParsedRecords


@dataclass(frozen=True)
class ImportSource:
    """与具体文件格式解耦的数据源元数据。"""

    source_type: str
    source_name: str | None
    source_hash: str | None
    source_format_version: int | None = None
    producer: str | None = None
    external_batch_id: str | None = None
    observed_at_utc: datetime | None = None


@dataclass(frozen=True)
class ImportStats:
    """一次合并或导入操作的完整统计。"""

    valid_rows: int
    groups: int
    members: int
    relations: int
    source_rows: int = 0
    skipped_rows: int = 0
    missing_user_id_rows: int = 0
    missing_group_id_rows: int = 0
    new_groups: int = 0
    updated_groups: int = 0
    new_members: int = 0
    new_relations: int = 0
    updated_relations: int = 0
    unchanged_relations: int = 0
    batch_id: int | None = None
    external_batch_id: str | None = None
    source_hash: str | None = None
    duplicate: bool = False
    duplicate_of: int | None = None
    search_index_status: str | None = None


def normalize_source_value(value: Any) -> str | None:
    """把来源字段转换为稳定的文本值。"""

    if value is None:
        return None
    if isinstance(value, datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        text = value.isoformat()
    else:
        text = str(value).strip()
    return text or None


_normalize_cell = normalize_source_value


def normalize_observed_at_utc(
    value: str | datetime | None,
) -> datetime | None:
    """把带时区的来源采集时间标准化为 UTC 无时区值。"""

    if value is None:
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.endswith(("Z", "z")):
            text = text[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError as exc:
            raise ValueError(
                "采集时间必须是带时区的 ISO 8601 时间"
            ) from exc
    if parsed.tzinfo is None:
        raise ValueError("采集时间必须包含时区")
    return parsed.astimezone(timezone.utc).replace(tzinfo=None)


def _validate_xlsx_path(filepath: str | Path) -> Path:
    path = Path(filepath)
    if not path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"仅支持 .xlsx 文件: {path}")
    return path


def calculate_file_hash(filepath: str | Path) -> str:
    """流式计算文件 SHA-256，用于识别重复数据源。"""

    path = _validate_xlsx_path(filepath)
    return calculate_path_hash(path)


def calculate_path_hash(filepath: str | Path) -> str:
    """计算任意已存在来源文件的 SHA-256。"""

    path = Path(filepath)
    if not path.is_file():
        raise FileNotFoundError(f"数据源文件不存在: {path}")
    digest = sha256()
    with path.open("rb") as source:
        while chunk := source.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def parse_xlsx_with_stats(filepath: str | Path) -> ParsedRecords:
    """读取全部工作表并返回记录及跳过原因统计。"""

    path = _validate_xlsx_path(filepath)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[Record] = []
    source_rows = 0
    skipped_rows = 0
    missing_user_id_rows = 0
    missing_group_id_rows = 0

    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration as exc:
                raise ValueError(f"工作表 {worksheet.title!r} 为空") from exc

            headers = {
                header: index
                for index, value in enumerate(header_row)
                if (header := _normalize_cell(value)) is not None
            }
            missing = sorted(set(REQUIRED_COLUMNS) - set(headers))
            if missing:
                missing_text = ", ".join(missing)
                raise ValueError(
                    f"工作表 {worksheet.title!r} 缺少必要列: {missing_text}"
                )

            for row in iterator:
                source_rows += 1
                record = {
                    column: (
                        _normalize_cell(row[headers[column]])
                        if column in headers and headers[column] < len(row)
                        else None
                    )
                    for column in SOURCE_COLUMNS
                }
                missing_user_id = not record["user_id"]
                missing_group_id = not record["group_id"]
                if missing_user_id:
                    missing_user_id_rows += 1
                if missing_group_id:
                    missing_group_id_rows += 1
                if missing_user_id or missing_group_id:
                    skipped_rows += 1
                    continue
                rows.append(record)
    finally:
        workbook.close()

    return ParsedRecords(
        rows=tuple(rows),
        source_rows=source_rows,
        skipped_rows=skipped_rows,
        missing_user_id_rows=missing_user_id_rows,
        missing_group_id_rows=missing_group_id_rows,
    )


def parse_xlsx(filepath: str | Path) -> list[Record]:
    """兼容接口：只返回工作簿中的有效记录。"""

    return list(parse_xlsx_with_stats(filepath).rows)


def _aggregate_rows(rows: Sequence[Record]):
    groups: dict[str, dict[str, str | None]] = {}
    members: set[str] = set()
    relations: dict[tuple[str, str], Record] = {}

    for record in rows:
        user_id = record["user_id"]
        group_id = record["group_id"]
        if user_id is None or group_id is None:
            continue

        group_name = record.get("group_name")
        if group_id not in groups:
            groups[group_id] = {
                "group_id": group_id,
                "group_name": group_name,
            }
        elif group_name is not None:
            groups[group_id]["group_name"] = group_name

        members.add(user_id)
        relation_key = (user_id, group_id)
        relation = relations.setdefault(
            relation_key,
            {
                "user_id": user_id,
                "group_id": group_id,
                **{field: None for field in RELATION_FIELDS},
            },
        )
        for field in RELATION_FIELDS:
            if (value := record.get(field)) is not None:
                relation[field] = value

    return groups, members, relations


def _existing_relations(session: SessionType) -> dict[tuple[str, str], dict]:
    columns = (
        MemberGroupInfo.user_id,
        MemberGroupInfo.group_id,
        *(getattr(MemberGroupInfo, field) for field in RELATION_FIELDS),
    )
    return {
        (row["user_id"], row["group_id"]): dict(row)
        for row in session.execute(select(*columns)).mappings()
    }


def import_to_db(
    rows: Sequence[Record],
    session: SessionType,
    *,
    batch_id: int | None = None,
) -> ImportStats:
    """使用非破坏性 upsert 合并数据，并按需记录观察批次。"""

    groups, members, relations = _aggregate_rows(rows)
    existing_groups = dict(
        session.execute(select(Group.group_id, Group.group_name)).all()
    )
    existing_members = set(session.scalars(select(Member.user_id)))
    existing_relations = _existing_relations(session)

    new_groups = sum(group_id not in existing_groups for group_id in groups)
    updated_groups = sum(
        group_id in existing_groups
        and values["group_name"] is not None
        and values["group_name"] != existing_groups[group_id]
        for group_id, values in groups.items()
    )
    new_members = sum(user_id not in existing_members for user_id in members)
    new_relations = sum(key not in existing_relations for key in relations)
    updated_relations = sum(
        key in existing_relations
        and any(
            values[field] is not None
            and values[field] != existing_relations[key][field]
            for field in RELATION_FIELDS
        )
        for key, values in relations.items()
    )
    search_relation_changed = any(
        key in existing_relations
        and any(
            values[field] is not None
            and values[field] != existing_relations[key][field]
            for field in SEARCH_CONTENT_RELATION_FIELDS
        )
        for key, values in relations.items()
    )
    unchanged_relations = len(relations) - new_relations - updated_relations

    if groups:
        group_insert = sqlite_insert(Group)
        group_upsert = group_insert.on_conflict_do_update(
            index_elements=[Group.group_id],
            set_={
                "group_name": func.coalesce(
                    group_insert.excluded.group_name,
                    Group.group_name,
                )
            },
        )
        session.execute(group_upsert, [groups[key] for key in sorted(groups)])

    if members:
        member_insert = sqlite_insert(Member).on_conflict_do_nothing(
            index_elements=[Member.user_id]
        )
        session.execute(
            member_insert,
            [{"user_id": user_id} for user_id in sorted(members)],
        )

    if relations:
        relation_insert = sqlite_insert(MemberGroupInfo)
        relation_upsert = relation_insert.on_conflict_do_update(
            index_elements=[
                MemberGroupInfo.user_id,
                MemberGroupInfo.group_id,
            ],
            set_={
                field: func.coalesce(
                    getattr(relation_insert.excluded, field),
                    getattr(MemberGroupInfo, field),
                )
                for field in RELATION_FIELDS
            },
        )
        session.execute(
            relation_upsert,
            [relations[key] for key in sorted(relations)],
        )

    if batch_id is not None and relations:
        observation_insert = sqlite_insert(RelationObservation)
        observation_upsert = observation_insert.on_conflict_do_update(
            index_elements=[
                RelationObservation.user_id,
                RelationObservation.group_id,
            ],
            set_={"last_seen_batch_id": batch_id},
        )
        session.execute(
            observation_upsert,
            [
                {
                    "user_id": user_id,
                    "group_id": group_id,
                    "first_seen_batch_id": batch_id,
                    "last_seen_batch_id": batch_id,
                }
                for user_id, group_id in sorted(relations)
            ],
        )

    session.flush()
    connection = session.connection()
    index_state = get_search_index_state(connection)
    search_content_changed = bool(
        new_relations or search_relation_changed or updated_groups
    )
    if search_content_changed or not (index_state and index_state["ready"]):
        search_index_status = rebuild_search_index(connection).status
    else:
        search_index_status = str(index_state["status"])

    return ImportStats(
        valid_rows=len(rows),
        groups=len(groups),
        members=len(members),
        relations=len(relations),
        source_rows=len(rows),
        new_groups=new_groups,
        updated_groups=updated_groups,
        new_members=new_members,
        new_relations=new_relations,
        updated_relations=updated_relations,
        unchanged_relations=unchanged_relations,
        batch_id=batch_id,
        search_index_status=search_index_status,
    )


class BatchIdentityConflictError(ValueError):
    """同一外部批次身份被用于不同内容。"""


def _find_duplicate(
    session: SessionType,
    source: ImportSource,
) -> ImportBatch | None:
    if source.external_batch_id is not None:
        batch = session.scalar(
            select(ImportBatch)
            .where(
                ImportBatch.producer == source.producer,
                ImportBatch.external_batch_id == source.external_batch_id,
            )
            .order_by(ImportBatch.id.desc())
            .limit(1)
        )
        if (
            batch is not None
            and batch.source_hash is not None
            and source.source_hash is not None
            and batch.source_hash != source.source_hash
        ):
            raise BatchIdentityConflictError(
                "同一 producer + batch_id 对应了不同内容"
            )
        return batch

    if source.source_hash is None:
        return None
    return session.scalar(
        select(ImportBatch)
        .where(
            ImportBatch.source_type == source.source_type,
            ImportBatch.source_hash == source.source_hash,
        )
        .order_by(ImportBatch.id.desc())
        .limit(1)
    )


def _stats_from_batch(
    batch: ImportBatch,
    *,
    duplicate: bool,
) -> ImportStats:
    return ImportStats(
        valid_rows=batch.valid_rows,
        groups=batch.unique_groups,
        members=batch.unique_members,
        relations=batch.unique_relations,
        source_rows=batch.source_rows,
        skipped_rows=batch.skipped_rows,
        missing_user_id_rows=batch.missing_user_id_rows,
        missing_group_id_rows=batch.missing_group_id_rows,
        new_groups=batch.new_groups,
        updated_groups=batch.updated_groups,
        new_members=batch.new_members,
        new_relations=batch.new_relations,
        updated_relations=batch.updated_relations,
        unchanged_relations=batch.unchanged_relations,
        batch_id=batch.id,
        external_batch_id=batch.external_batch_id,
        source_hash=batch.source_hash,
        duplicate=duplicate,
        duplicate_of=batch.id if duplicate else batch.duplicate_of_id,
    )


def _print_import_stats(stats: ImportStats) -> None:
    if stats.duplicate:
        safe_print(
            f"数据源已由批次 #{stats.duplicate_of} 导入，"
            "本次未重复写入；使用 --force 可强制处理。"
        )
        if stats.search_index_status != "ready":
            safe_print("搜索索引未就绪，查询将自动回退到 LIKE。")
        return

    safe_print(
        f"导入完成（批次 #{stats.batch_id}）: "
        f"{stats.groups} 个群组, "
        f"{stats.members} 个成员, "
        f"{stats.relations} 条成员-群组关系"
    )
    safe_print(
        "合并统计: "
        f"新增群组 {stats.new_groups}, "
        f"更新群组 {stats.updated_groups}, "
        f"新增成员 {stats.new_members}, "
        f"新增关系 {stats.new_relations}, "
        f"更新关系 {stats.updated_relations}, "
        f"未变化关系 {stats.unchanged_relations}"
    )
    if stats.skipped_rows:
        safe_print(
            "跳过行: "
            f"{stats.skipped_rows} "
            f"（缺少 user_id: {stats.missing_user_id_rows}, "
            f"缺少 group_id: {stats.missing_group_id_rows}）"
        )
    if stats.search_index_status != "ready":
        safe_print("搜索索引未就绪，查询将自动回退到 LIKE。")


def import_parsed_records(
    parsed: ParsedRecords,
    source: ImportSource,
    db_path: str | Path = DB_PATH,
    *,
    force: bool = False,
) -> ImportStats:
    """把任意已解析标准批次合并到数据库。"""

    if not source.source_type.strip():
        raise ValueError("数据源类型不能为空")
    if source.source_format_version is not None:
        if source.source_format_version < 1:
            raise ValueError("数据源格式版本必须大于 0")
    source_type = source.source_type.strip()
    producer = source.producer.strip() if source.producer is not None else None
    external_batch_id = (
        source.external_batch_id.strip()
        if source.external_batch_id is not None
        else None
    )
    if external_batch_id and not producer:
        raise ValueError("外部 batch_id 必须与 producer 一同提供")
    if external_batch_id and len(external_batch_id) > 128:
        raise ValueError("外部 batch_id 不能超过 128 个字符")
    source = replace(
        source,
        source_type=source_type,
        producer=producer or None,
        external_batch_id=external_batch_id or None,
    )

    engine, Session = init_db(db_path, create=True)
    try:
        with Session() as session:
            duplicate_batch = _find_duplicate(session, source)
            if (
                duplicate_batch is not None
                and force
                and source.external_batch_id is not None
            ):
                raise BatchIdentityConflictError(
                    "带 batch_id 的批次不能强制重复处理；请使用新的 batch_id"
                )
            if duplicate_batch is not None and not force:
                state = get_search_index_state(session.connection())
                stats = replace(
                    _stats_from_batch(duplicate_batch, duplicate=True),
                    search_index_status=(
                        str(state["status"]) if state else "stale"
                    ),
                )
                _print_import_stats(stats)
                return stats

        with Session.begin() as session:
            duplicate_batch = _find_duplicate(session, source)
            batch = ImportBatch(
                source_type=source.source_type,
                source_name=source.source_name,
                source_hash=source.source_hash,
                source_format_version=source.source_format_version,
                producer=source.producer,
                external_batch_id=source.external_batch_id,
                observed_at_utc=source.observed_at_utc,
                imported_at_utc=datetime.now(timezone.utc).replace(tzinfo=None),
                forced=force,
                duplicate_of_id=duplicate_batch.id
                if duplicate_batch is not None
                else None,
                source_rows=parsed.source_rows,
                valid_rows=len(parsed.rows),
                skipped_rows=parsed.skipped_rows,
                missing_user_id_rows=parsed.missing_user_id_rows,
                missing_group_id_rows=parsed.missing_group_id_rows,
            )
            session.add(batch)
            session.flush()

            merge_stats = import_to_db(
                parsed.rows,
                session,
                batch_id=batch.id,
            )
            batch.unique_groups = merge_stats.groups
            batch.unique_members = merge_stats.members
            batch.unique_relations = merge_stats.relations
            batch.new_groups = merge_stats.new_groups
            batch.updated_groups = merge_stats.updated_groups
            batch.new_members = merge_stats.new_members
            batch.new_relations = merge_stats.new_relations
            batch.updated_relations = merge_stats.updated_relations
            batch.unchanged_relations = merge_stats.unchanged_relations

            stats = replace(
                merge_stats,
                source_rows=parsed.source_rows,
                skipped_rows=parsed.skipped_rows,
                missing_user_id_rows=parsed.missing_user_id_rows,
                missing_group_id_rows=parsed.missing_group_id_rows,
                source_hash=source.source_hash,
                external_batch_id=source.external_batch_id,
                duplicate_of=batch.duplicate_of_id,
            )

        _print_import_stats(stats)
        return stats
    finally:
        engine.dispose()


def import_xlsx(
    filepath: str | Path,
    db_path: str | Path = DB_PATH,
    *,
    force: bool = False,
    producer: str | None = None,
    observed_at_utc: str | datetime | None = None,
) -> ImportStats:
    """解析并导入 xlsx；相同文件默认只处理一次。"""

    path = _validate_xlsx_path(filepath)
    safe_print(f"正在解析: {path}")
    parsed = parse_xlsx_with_stats(path)
    safe_print(
        f"读取 {parsed.source_rows} 行，"
        f"有效 {len(parsed.rows)} 行，"
        f"跳过 {parsed.skipped_rows} 行"
    )
    normalized_producer = str(producer).strip() if producer is not None else None
    source = ImportSource(
        source_type="xlsx",
        source_name=path.name,
        source_hash=calculate_file_hash(path),
        source_format_version=1,
        producer=normalized_producer or None,
        observed_at_utc=normalize_observed_at_utc(observed_at_utc),
    )
    return import_parsed_records(parsed, source, db_path, force=force)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="导入群成员 xlsx 数据")
    parser.add_argument("xlsx_path", help="xlsx 文件路径")
    parser.add_argument("--db", default=DB_PATH, help="SQLite 数据库路径")
    parser.add_argument(
        "--force",
        action="store_true",
        help="即使文件哈希已导入也强制处理",
    )
    parser.add_argument("--producer", help="数据生产方，例如 astrbot")
    parser.add_argument(
        "--observed-at",
        help="带时区的 ISO 8601 数据采集时间",
    )
    arguments = parser.parse_args()
    import_xlsx(
        arguments.xlsx_path,
        arguments.db,
        force=arguments.force,
        producer=arguments.producer,
        observed_at_utc=arguments.observed_at,
    )
