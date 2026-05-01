"""从 xlsx 文件导入数据到数据库"""

import os
from openpyxl import load_workbook
from sqlalchemy.orm import Session as SessionType
from config import DB_PATH, REQUIRED_COLUMNS
from models import Group, Member, MemberGroupInfo, init_db


def parse_xlsx(filepath: str) -> list[dict]:
    """
    解析 xlsx 文件，返回所有分页中的有效行数据。
    每行返回一个 dict，只包含 REQUIRED_COLUMNS 中的字段。
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = {str(cell).strip(): idx for idx, cell in enumerate(row) if cell is not None}
                continue

            if headers is None:
                continue

            record = {}
            for col_name in REQUIRED_COLUMNS:
                idx = headers.get(col_name)
                if idx is not None and idx < len(row):
                    value = row[idx]
                    record[col_name] = str(value).strip() if value is not None else None
                else:
                    record[col_name] = None

            if record.get("user_id"):
                rows.append(record)

    wb.close()
    return rows


def import_to_db(rows: list[dict], session: SessionType):
    """
    将解析出的行数据导入数据库。
    - group_id 与 group_name 一对一绑定
    - user_id 作为唯一标识，相同 user_id 合并
    - 同一 user_id + group_id 下的 nickname/card/join_time/last_sent_time/title 为一个绑定组
    """
    for record in rows:
        user_id = record["user_id"]
        group_id = record.get("group_id")
        group_name = record.get("group_name")

        if group_id:
            group = session.get(Group, group_id)
            if group is None:
                group = Group(group_id=group_id, group_name=group_name)
                session.add(group)
            elif group_name and group.group_name is None:
                group.group_name = group_name

        member = session.get(Member, user_id)
        if member is None:
            member = Member(user_id=user_id)
            session.add(member)

        if group_id:
            info = session.get(MemberGroupInfo, (user_id, group_id))
            if info is None:
                info = MemberGroupInfo(
                    user_id=user_id,
                    group_id=group_id,
                    nickname=record.get("nickname"),
                    card=record.get("card"),
                    join_time=record.get("join_time"),
                    last_sent_time=record.get("last_sent_time"),
                    title=record.get("title"),
                )
                session.add(info)
            else:
                for field in ("nickname", "card", "join_time", "last_sent_time", "title"):
                    new_val = record.get(field)
                    if new_val is not None:
                        setattr(info, field, new_val)

    session.commit()


def import_xlsx(filepath: str, db_path: str = DB_PATH):
    """完整导入流程：解析 xlsx -> 写入数据库"""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    _, Session = init_db(db_path)
    session = Session()

    try:
        print(f"正在解析: {filepath}")
        rows = parse_xlsx(filepath)
        print(f"解析到 {len(rows)} 条有效记录")

        print("正在导入数据库...")
        import_to_db(rows, session)
        print("导入完成!")
    except Exception as e:
        session.rollback()
        print(f"导入失败: {e}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python importer.py <xlsx文件路径> [数据库路径]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    db = sys.argv[2] if len(sys.argv) > 2 else DB_PATH
    import_xlsx(xlsx_path, db)
