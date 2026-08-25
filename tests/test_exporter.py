"""搜索结果导出测试。"""

import csv
import json

import pytest
from openpyxl import load_workbook

from social_database.config import REQUIRED_COLUMNS
from social_database.exporter import export_search_results, format_export_result
from social_database.importer import import_xlsx


def record(**values):
    result = {column: None for column in REQUIRED_COLUMNS}
    result.update(values)
    return result


def test_export_json_csv_and_xlsx(tmp_path, workbook_factory):
    workbook = workbook_factory(
        tmp_path / "source.xlsx",
        {
            "Members": [
                record(
                    group_id="g-1",
                    user_id="u-1",
                    nickname="Alice",
                    group_name="Alpha Group",
                ),
                record(
                    group_id="g-2",
                    user_id="u-1",
                    card="Alice Card",
                    group_name="Second Group",
                ),
                record(
                    group_id="g-3",
                    user_id="u-2",
                    nickname="Bob",
                    group_name="Third Group",
                ),
            ]
        },
    )
    database = tmp_path / "members.db"
    import_xlsx(workbook, database)

    json_path = tmp_path / "exports" / "alice.json"
    csv_path = tmp_path / "exports" / "alice.csv"
    xlsx_path = tmp_path / "exports" / "alice.xlsx"

    json_result = export_search_results(
        "Alice",
        json_path,
        database,
        field="nickname",
    )
    csv_result = export_search_results(
        "Alice",
        csv_path,
        database,
        field="nickname",
    )
    xlsx_result = export_search_results(
        "Alice",
        xlsx_path,
        database,
        field="nickname",
    )

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["count"] == 1
    assert len(payload["results"][0]["groups"]) == 2
    assert payload["results"][0]["groups"][0]["first_seen_batch_id"] == 1

    with csv_path.open(encoding="utf-8-sig", newline="") as source:
        csv_rows = list(csv.DictReader(source))
    assert len(csv_rows) == 2
    assert {row["group_id"] for row in csv_rows} == {"g-1", "g-2"}
    assert {row["first_seen_batch_id"] for row in csv_rows} == {"1"}
    assert "area" in csv_rows[0]

    exported_workbook = load_workbook(
        xlsx_path,
        read_only=True,
        data_only=True,
    )
    try:
        xlsx_rows = list(exported_workbook.active.iter_rows(values_only=True))
    finally:
        exported_workbook.close()
    assert xlsx_rows[0][0:3] == ("user_id", "group_id", "group_name")
    assert "last_seen_at_utc" in xlsx_rows[0]
    assert len(xlsx_rows) == 3

    for result in (json_result, csv_result, xlsx_result):
        assert result["users"] == 1
        assert result["relations"] == 2
    assert "1 个用户、2 条成员-群关系" in format_export_result(json_result)

    with pytest.raises(FileExistsError, match="导出文件已存在"):
        export_search_results("Alice", json_path, database)
    overwritten = export_search_results(
        "Alice",
        json_path,
        database,
        overwrite=True,
    )
    assert overwritten["format"] == "json"
    assert not list(json_path.parent.glob(".*.tmp"))


def test_export_rejects_unknown_extension(tmp_path):
    with pytest.raises(ValueError, match="不支持的导出格式"):
        export_search_results(
            "Alice",
            tmp_path / "results.txt",
            tmp_path / "missing.db",
        )
